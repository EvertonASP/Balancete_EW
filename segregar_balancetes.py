import os
import unicodedata
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side, numbers
from datetime import datetime
import tkinter as tk
import calendar
from tkinter import ttk

# Função para limpar nome do arquivo
def limpar_nome(nome):
    nome = unicodedata.normalize('NFKD', nome).encode('ASCII', 'ignore').decode('ASCII')
    nome = ''.join(c if c.isalnum() or c in ['_', '-'] else '_' for c in nome)
    return nome.strip('_')

# Função para normalizar nome da empresa
def normalizar_nome_empresa(nome):
    nome = unicodedata.normalize('NFKD', nome).encode('ASCII', 'ignore').decode('ASCII')
    nome = nome.strip().upper()
    return nome

# Função para ajustar largura, filtros, congelamento, formatação, bordas e remover linhas de grade
def ajustar_largura_colunas(caminho_arquivo):
    wb = load_workbook(caminho_arquivo)
    ws = wb.active

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Ajusta largura com base no maior conteúdo
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        ws.column_dimensions[col_letter].width = max_length + 3  # margem extra

    # Estilo do cabeçalho: fundo azul escuro, texto branco, negrito e centralizado
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Preenche todas as células com fundo branco para ocultar linhas de grade
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Formatação: valores numéricos alinhados à direita e com duas casas decimais
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = white_fill
            cell.border = thin_border
            if isinstance(cell.value, (int, float)):
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    # Adiciona filtros automáticos
    ws.auto_filter.ref = ws.dimensions
    # Congela a primeira linha
    ws.freeze_panes = "A2"

    wb.save(caminho_arquivo)

# ✅ Interface gráfica para selecionar mês
def selecionar_mes():
    janela = tk.Tk()
    janela.title("Selecione o mês")
    janela.geometry("300x150")

    tk.Label(janela, text="Escolha o mês:", font=("Arial", 12)).pack(pady=10)

    meses_lista = ["janeiro", "fevereiro", "março", "abril", "maio", "junho",
                   "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"]

    mes_var = tk.StringVar()
    combo = ttk.Combobox(janela, textvariable=mes_var, values=meses_lista, state="readonly", font=("Arial", 12))
    combo.pack(pady=10)
    combo.current(8)  # padrão: setembro

    def confirmar():
        janela.destroy()

    tk.Button(janela, text="Confirmar", command=confirmar, font=("Arial", 12)).pack(pady=10)

    janela.mainloop()
    return mes_var.get()

# Usa a função para obter o mês
mes_input = selecionar_mes()
ano = 2025
meses = {
    "janeiro": 1, "fevereiro": 2, "março": 3, "abril": 4, "maio": 5, "junho": 6,
    "julho": 7, "agosto": 8, "setembro": 9, "outubro": 10, "novembro": 11, "dezembro": 12
}
if mes_input not in meses:
    raise ValueError("Mês inválido!")

# Obtém o último dia do mês selecionado
ultimo_dia = calendar.monthrange(ano, meses[mes_input])[1]

# Formata a data com o último dia do mês
data_mes = datetime(ano, meses[mes_input], ultimo_dia).strftime("%d/%m/%Y")

# ✅ Lê o arquivo de mapeamento (Excel)
arquivo_depara = r"C:\\Users\\everton.pinto\\Grupo Ultra\\Grupo Ultra - IRPJ e CSLL - IRPJ CSLL\\13. Consultas e Projetos\\10.0 Relação empresas Easy Way\\Automação Balancetes EW\\depara_empresas.xlsx"
df_depara = pd.read_excel(arquivo_depara)
mapa_empresas = {normalizar_nome_empresa(row["EMPRESA"]): row["CODIGO"] for _, row in df_depara.iterrows()}

# Função para salvar bloco
def salvar_bloco(dados, filial):
    df = pd.DataFrame(dados, columns=cabecalho)

    # Adiciona colunas fixas
    for nome_col, valor_padrao in reversed(colunas_fixas):
        pos = df.columns.get_loc("SALDO_ANTERIOR") if "SALDO_ANTERIOR" in df.columns else len(df.columns)
        df.insert(pos, nome_col, valor_padrao)

    # Remove colunas indesejadas
    for col_excluir in ["ID", "REF.", "SALDO_ANTERIOR"]:
        if col_excluir in df.columns:
            df.drop(columns=[col_excluir], inplace=True)

    # ✅ Substitui valores da coluna COD_EMPRESA pelo código do dicionário
    if "COD_EMPRESA" in df.columns:
        codigo_empresa = mapa_empresas.get(normalizar_nome_empresa(filial), "N/A")
        df["COD_EMPRESA"] = codigo_empresa

    # ✅ Adiciona DATA_REFERENCIA
    if "DATA_REFERENCIA" not in df.columns:
        df.insert(2, "DATA_REFERENCIA", data_mes)

    # Salva Excel
    nome_limpo = limpar_nome(filial)
    
    # Código da empresa e nome formatado
    codigo_empresa = mapa_empresas.get(normalizar_nome_empresa(filial), "CODIGO")
    nome_empresa = limpar_nome(filial)

    # Formata mês e ano no padrão MM.AA
    mes_ano_formatado = f"{meses[mes_input]:02d}.{str(ano)[-2:]}"

    # Novo padrão do nome do arquivo
    nome_arquivo = f"{codigo_empresa}_{nome_empresa}_BALANCETE EW_{mes_ano_formatado}.xlsx"

    # Caminho completo
    caminho_arquivo = os.path.join(pasta_saida, nome_arquivo)
    
    with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Balancete")

    ajustar_largura_colunas(caminho_arquivo)

# Caminhos
arquivo_origem = r"C:\\Users\\everton.pinto\\Grupo Ultra\\Grupo Ultra - IRPJ e CSLL - IRPJ CSLL\\13. Consultas e Projetos\\10.0 Relação empresas Easy Way\\Automação Balancetes EW\\Balancete MLN-WP-Consolidado_v2.xlsm"

# Defina a pasta de saída como um caminho absoluto
pasta_saida = r"C:\\Users\\everton.pinto\\Grupo Ultra\\Grupo Ultra - IRPJ e CSLL - IRPJ CSLL\\13. Consultas e Projetos\\10.0 Relação empresas Easy Way\\Automação Balancetes EW\\Balancetes_Saida" 

os.makedirs(pasta_saida, exist_ok=True)

# Carrega planilha
wb = load_workbook(filename=arquivo_origem, data_only=True)
ws = wb["Balancete MLN_WebPosto"]

# Cabeçalho fixo
cabecalho = [ws.cell(row=22, column=col).value for col in range(2, 11)]

# Colunas fixas
colunas_fixas = [
    ("TOTAL_SALDO_INICIAL", ""),
    ("IND_NATUREZA_1", ""),
    ("DESCRICAO_SINTETICA", ""),
    ("COD_CONTA_SINTETICA", ""),
    ("IND_NATUREZA", ""),
]

# Processa blocos
dados_bloco = []
nome_filial = None

for row in range(23, ws.max_row + 1):
    valor_b = ws.cell(row=row, column=2).value
    valor_c = ws.cell(row=row, column=3).value

    if valor_b and valor_c:
        if nome_filial and valor_b != nome_filial:
            salvar_bloco(dados_bloco, nome_filial)
            dados_bloco = []
        nome_filial = valor_b
        dados_bloco.append([ws.cell(row=row, column=col).value for col in range(2, 11)])
    else:
        if nome_filial and dados_bloco:
            salvar_bloco(dados_bloco, nome_filial)
            dados_bloco = []
            nome_filial = None

if nome_filial and dados_bloco:
    salvar_bloco(dados_bloco, nome_filial)

print("✅ Arquivos gerados com sucesso!")